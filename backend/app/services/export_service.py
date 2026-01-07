from io import BytesIO
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def export_to_pdf(expenses, user, start_date, end_date):
    """Exporta gastos a PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        alignment=1
    )
    title = Paragraph("Reporte de Gastos", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Información del usuario y período
    info_style = styles['Normal']
    info = Paragraph(
        f"<b>Usuario:</b> {user.username}<br/>"
        f"<b>Período:</b> {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}<br/>"
        f"<b>Generado:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        info_style
    )
    elements.append(info)
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabla de gastos
    if expenses:
        data = [['Fecha', 'Categoría', 'Monto', 'Nota']]
        total = 0
        
        for expense in expenses:
            data.append([
                expense.date.strftime('%d/%m/%Y'),
                expense.category.name if hasattr(expense, 'category') else 'N/A',
                f"{user.currency} {expense.amount:.2f}",
                expense.note[:30] if expense.note else '-'
            ])
            total += expense.amount
        
        data.append(['', '', '', ''])
        data.append(['', 'TOTAL', f"{user.currency} {total:.2f}", ''])
        
        table = Table(data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 2.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -3), colors.beige),
            ('GRID', (0, 0), (-1, -3), 1, colors.black),
            ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -2), (-1, -1), colors.HexColor('#ecf0f1')),
        ]))
        elements.append(table)
    else:
        no_data = Paragraph("No hay gastos registrados en este período.", styles['Normal'])
        elements.append(no_data)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def export_to_excel(expenses_with_categories, start_date, end_date):
    """Exporta gastos a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"
    
    # Headers
    headers = ['Fecha', 'Hora', 'Categoría', 'Monto', 'Moneda', 'Nota']
    ws.append(headers)
    
    # Estilo de headers
    header_font = Font(bold=True, size=12)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    total = 0
    for expense, category_name in expenses_with_categories:
        ws.append([
            expense.date.strftime('%d/%m/%Y'),
            expense.date.strftime('%H:%M:%S'),
            category_name,
            expense.amount,
            expense.currency,
            expense.note if expense.note else ''
        ])
        total += expense.amount
    
    # Fila de total
    ws.append([])
    total_row = ws.max_row
    ws[f'A{total_row}'] = 'TOTAL'
    ws[f'D{total_row}'] = total
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'D{total_row}'].font = Font(bold=True)
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 40
    
    # Guardar en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()